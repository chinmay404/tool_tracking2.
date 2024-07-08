# import uuid
# import qrcode

# def generate_qr_code(uuid_value):
#     qr = qrcode.QRCode(
#         version=1,
#         error_correction=qrcode.constants.ERROR_CORRECT_L,
#         box_size=10,
#         border=4,
#     )
#     qr.add_data(str(uuid_value))
#     qr.make(fit=True)

#     qr_img = qr.make_image(fill_color="black", back_color="white")
#     return qr_img

# if __name__ == "__main__":
#     uuid_value = uuid.uuid4()
#     print(uuid_value)
#     qr_code = generate_qr_code(uuid_value)

#     qr_code_path = "output_qr_code.png"
#     qr_code.save(qr_code_path)
#     print(f"QR Code saved as '{qr_code_path}'")


# import uuid
# import qrcode

# def generate_qr_code(uuid_value):
#     qr = qrcode.QRCode(
#         version=1,
#         error_correction=qrcode.constants.ERROR_CORRECT_L,
#         box_size=10,
#         border=4,
#     )
#     qr.add_data(str(uuid_value))
#     qr.make(fit=True)

#     qr_img = qr.make_image(fill_color="black", back_color="white")
#     return qr_img

# if __name__ == "__main__":
#     uuid_value = uuid.uuid4()
#     uuid_str = str(uuid_value).replace("-", "")  # Convert UUID to string without hyphens
#     qr_code = generate_qr_code(uuid_str)  # Use the UUID string to generate QR code

#     qr_code_path = f"{uuid_str}_qr_code.png"  # Use the UUID string in the filename
#     qr_code.save(qr_code_path)
#     print(f"QR Code saved as '{qr_code_path}'")


import uuid
import qrcode
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
import os


def generate_qr_code(uuid_value):
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(str(uuid_value))
    qr.make(fit=True)

    qr_img = qr.make_image(fill_color="black", back_color="white")
    return qr_img


if __name__ == "__main__":
    num_qrcodes = int(input("Enter the number of QR codes to generate: "))

    wb = Workbook()
    ws = wb.active

    for i in range(num_qrcodes):
        uuid_value = uuid.uuid4()
        uuid_str = str(uuid_value).replace("-", "")

        qr_code = generate_qr_code(uuid_str)
        qr_code_path = f"{uuid_str}.png"
        qr_code.save(qr_code_path)

        img = ExcelImage(qr_code_path)

        col = (i % 2) * 2 + 1
        row = (i // 2) * 15 + 1 

        col_letter = get_column_letter(col)
        merge_range = f"{col_letter}{row}:{get_column_letter(col + 1)}{row + 14}"

        ws.merge_cells(merge_range)
        cell = ws.cell(row=row, column=col)
        ws.add_image(img, f"{cell.coordinate}")

        ws.cell(row=row + 15, column=col, value=uuid_str)

    excel_filename = "qrcodes.xlsx"
    wb.save(excel_filename)
    print(f'QR codes saved to "{excel_filename}"')

    # Clean up generated QR code image files
    for qr_code_path in os.listdir():
        if qr_code_path.endswith(".png"):
            os.remove(qr_code_path)


